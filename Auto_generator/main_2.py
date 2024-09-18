import openpyxl
from jinja2 import Environment, FileSystemLoader
from datetime import datetime

# Path to your Excel file
excel_file = 'Team_Info.xlsx'

# Load the Excel workbook
workbook = openpyxl.load_workbook(excel_file)
sheet = workbook.active

# Create a Jinja2 environment
env = Environment(loader=FileSystemLoader('.'))
template = env.get_template('template_2.html')

# Initialize variables
html_output = ''
data_rows = []

# Loop through each row in the sheet, starting from row 2
for row in range(2, sheet.max_row + 1):
    Name = sheet.cell(row=row, column=1).value
    Position = sheet.cell(row=row, column=2).value
    Cohort_year = sheet.cell(row=row, column=3).value
    Research_Interest = sheet.cell(row=row, column=4).value  # Assuming year is in the fourth column
    Personal_Website = sheet.cell(row=row, column=5).value
    Image = "images/Team_Profile_Pic/"+str(row)+".jpg"

    # Format the year_cell if it's a datetime object
    if isinstance(Cohort_year, datetime):
        formatted_year = Cohort_year.strftime("%B %Y")  # Format to "Month Year" (e.g., "June 2024")
    else:
        formatted_year = Cohort_year  # Fallback if it's not a datetime object

    # Store the data in a list for later use
    data_rows.append((Name, Position, Cohort_year, Research_Interest, Personal_Website, Image))

    # Generate the HTML segment using the template
    html_segment = template.render(
        Name=Name,
        Position=Position,
        Cohort_year=Cohort_year,
        Research_Interest=Research_Interest,
        Personal_Website=Personal_Website,
        Image_Link=Image# Pass the formatted year to the template
    )

    # Append the HTML segment to the output
    if row == sheet.max_row:
        html_output += html_segment  # Last row, no need for extra div closing and opening
    else:
        html_output += html_segment + '\n'  # Add new div only if not the last row

# Print the top 5 data rows
print("Top 5 data rows:")
for i in range(min(5, len(data_rows))):  # Ensure we don't exceed the number of rows
    print(data_rows[i])

# Save the generated HTML to a text file
output_file_path = 'output_2.txt'
with open(output_file_path, 'w', encoding='utf-8') as file:
    file.write(html_output.strip())  # Write the stripped HTML output
    file.write('\n\n')  # Add two new lines
    file.write('This is a new line after the HTML output.')  # Write additional text

print(f"Generated HTML has been saved to {output_file_path}")