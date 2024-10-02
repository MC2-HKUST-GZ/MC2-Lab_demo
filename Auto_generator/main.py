import openpyxl
from jinja2 import Environment, FileSystemLoader
from datetime import datetime

# Path to your Excel file
excel_file = 'Auto_generator\Publication.xlsx'

# Load the Excel workbook
workbook = openpyxl.load_workbook(excel_file, data_only=True)
sheet = workbook.active

# Create a Jinja2 environment
env = Environment(loader=FileSystemLoader('.'))
template = env.get_template('Auto_generator/template.html')

# Initialize variables
html_output = ''


# Loop through each row in the sheet, starting from row 2
for row in range(2, sheet.max_row + 1):
    title = sheet.cell(row=row, column=1).value
    authors = sheet.cell(row=row, column=2).value
    year_cell = sheet.cell(row=row, column=3).value
    publication_info = sheet.cell(row=row, column=4).value  # Assuming year is in the fourth column
    doi_info = sheet.cell(row=row, column=6).value
    img_info = sheet.cell(row=row, column=7).value

    # Format the year_cell if it's a datetime object
    if isinstance(year_cell, datetime):
        formatted_year = year_cell.strftime("%B %Y")  # Format to "Month Year" (e.g., "June 2024")
    else:
        formatted_year = year_cell  # Fallback if it's not a datetime object

    # Generate the HTML segment using the template
    html_segment = template.render(
        title= title,
        authors= authors,
        publication_info= publication_info,
        year= formatted_year,  # Pass the formatted year to the template
        doi = "https://doi.org/"+str(doi_info),
        img = "images/Publication_png/"+str(img_info)+".png"
    )

    # Append the HTML segment to the output
    if row == sheet.max_row:
        html_output += html_segment  # Last row, no need for extra div closing and opening
    else:
        html_output += html_segment + '\n</div>\n<div class="row">\n'  # Add new div only if not the last row

# Save the generated HTML to a text file
output_file_path = 'output_3.txt'
with open(output_file_path, 'w', encoding='utf-8') as file:
    file.write(html_output.strip())  # Write the stripped HTML output
    file.write('\n\n')  # Add two new lines
    file.write('This is a new line after the HTML output.')  # Write additional text

print(f"Generated HTML has been saved to {output_file_path}")